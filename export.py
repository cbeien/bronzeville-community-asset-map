"""
export.py — All export functions for the Bronzeville Community Asset Map.

Produces four deliverables:
  1. bronzeville_map.html   — self-contained Folium interactive map
  2. bronzeville_assets.xlsx — multi-sheet Excel workbook
  3. data.js                — JavaScript data file for the Leaflet website
  4. index.html             — standalone Leaflet.js website (data embedded inline)
  5. bronzeville_summary.html — formatted HTML summary tables (bonus)
"""

from __future__ import annotations

import json
import math
import logging
import folium
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import (
    ZONES, MARKER_STYLES, ASSET_CATEGORIES,
    OUTPUT_MAP_HTML, OUTPUT_EXCEL, OUTPUT_DATA_JS,
    OUTPUT_INDEX_HTML, OUTPUT_SUMMARY,
    BRONZEVILLE_CENTER,
)

logger = logging.getLogger(__name__)

# Shared column order for asset sheets
ASSET_COLS = ["Name", "Address", "Type", "category", "latitude", "longitude", "source"]


# ═══════════════════════════════════════════════════════════════════════════════
# 1. Folium HTML Map
# ═══════════════════════════════════════════════════════════════════════════════

def export_html_map(m: folium.Map, path: str = OUTPUT_MAP_HTML) -> None:
    """Save the assembled Folium map as a self-contained HTML file."""
    m.save(path)
    print(f"  Folium map  → {path}")


# ═══════════════════════════════════════════════════════════════════════════════
# 2. Excel Workbook
# ═══════════════════════════════════════════════════════════════════════════════

# Header colours per sheet (hex, no '#')
_SHEET_COLORS = {
    "Landmarks":             "1F4E79",
    "Restaurants":           "7B241C",
    "Businesses":            "1A5C3C",
    "Schools":               "4A235A",
    "Houses of Worship":     "784212",
    "Parks":                 "145A32",
    "Healthcare":            "641E16",
    "Cultural Institutions": "0E6655",
    "Social Services":       "922B21",
    "Zone Summary":          "2C3E50",
    # Transportation sheets
    "CTA Rail Stations":     "D35400",
    "CTA Rail Lines":        "D35400",
    "CTA Bus Stops":         "7D3C98",
    "CTA Bus Routes":        "7D3C98",
    "Metra Stations":        "1A5276",
    "Metra Lines":           "1A5276",
}


def _style_header(ws, hex_color: str = "1F4E79") -> None:
    """Bold white text on a dark background for the first (header) row."""
    fill   = PatternFill("solid", fgColor=hex_color)
    font   = Font(bold=True, color="FFFFFF")
    align  = Alignment(horizontal="center", vertical="center")
    border = Border(bottom=Side(style="medium", color="FFFFFF"))
    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = align
        cell.border    = border


def _auto_width(ws) -> None:
    """Fit each column to its widest value (capped at 60 chars)."""
    for col in ws.columns:
        max_len = max(
            (len(str(c.value or "")) for c in col),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


def _write_df_sheet(wb: Workbook,
                    title: str,
                    df: pd.DataFrame,
                    cols: list[str]) -> None:
    """Write selected columns of df to a new sheet in wb."""
    ws = wb.create_sheet(title=title[:31])   # Excel tab names max 31 chars

    available = [c for c in cols if c in df.columns]
    ws.append(available)

    for _, row in df.iterrows():
        ws.append([str(row.get(c, ""))[:32767] for c in available])

    _style_header(ws, _SHEET_COLORS.get(title, "2C3E50"))
    _auto_width(ws)
    ws.freeze_panes = "A2"


def _write_zone_summary(wb: Workbook,
                        all_assets: pd.DataFrame) -> pd.DataFrame:
    """
    Count assets per zone per category and write a summary sheet.
    Returns the summary as a DataFrame for console display.
    """
    ws = wb.create_sheet(title="Zone Summary")

    # Build header dynamically from ASSET_CATEGORIES
    header = ["Zone", "Zone Label"] + [c.title() for c in ASSET_CATEGORIES] + ["Total"]
    ws.append(header)

    rows = []
    all_assets = all_assets.copy()
    all_assets["latitude"] = pd.to_numeric(all_assets.get("latitude"), errors="coerce")

    for zone_key, zone in ZONES.items():
        mask  = (
            (all_assets["latitude"] >= zone["lat_min"]) &
            (all_assets["latitude"] <  zone["lat_max"])
        )
        zone_df = all_assets[mask]

        counts = {
            cat: int((zone_df.get("category", pd.Series()) == cat).sum())
            for cat in ASSET_CATEGORIES
        }
        total = sum(counts.values())
        row   = [zone_key, zone["label"]] + list(counts.values()) + [total]
        ws.append(row)
        rows.append({"Zone": zone_key, "Label": zone["label"],
                     **counts, "Total": total})

    _style_header(ws, _SHEET_COLORS["Zone Summary"])
    _auto_width(ws)
    ws.freeze_panes = "A2"

    return pd.DataFrame(rows)


def export_excel(assets: dict[str, pd.DataFrame],
                 transport: dict[str, pd.DataFrame],
                 path: str = OUTPUT_EXCEL) -> pd.DataFrame:
    """
    Write the multi-sheet Excel workbook.

    Sheets produced
    ---------------
    Asset sheets (9):   Landmarks, Restaurants, Businesses, Schools,
                        Houses of Worship, Parks, Healthcare,
                        Cultural Institutions, Social Services
    Zone Summary (1):   Asset counts per zone per category
    Transport sheets (6): CTA Rail Stations, CTA Rail Lines,
                           CTA Bus Stops, CTA Bus Routes,
                           Metra Stations, Metra Lines

    Returns the zone summary DataFrame (also printed in main.py).
    """
    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    asset_sheet_map = {
        "landmarks":   "Landmarks",
        "restaurants": "Restaurants",
        "businesses":  "Businesses",
        "education":   "Education & Culture",
        "worship":     "Houses of Worship",
        "parks":       "Parks",
        "health":      "Health & Social Services",
    }

    for key, sheet_name in asset_sheet_map.items():
        _write_df_sheet(wb, sheet_name, assets.get(key, pd.DataFrame()), ASSET_COLS)

    # Zone summary — needs a combined frame with a 'category' column
    all_assets = pd.concat(list(assets.values()), ignore_index=True)
    zone_df    = _write_zone_summary(wb, all_assets)

    # Transportation sheets
    transport_sheet_map = {
        "cta_rail_stations": "CTA Rail Stations",
        "cta_rail_lines":    "CTA Rail Lines",
        "cta_bus_stops":     "CTA Bus Stops",
        "cta_bus_routes":    "CTA Bus Routes",
        "metra_stations":    "Metra Stations",
        "metra_lines":       "Metra Lines",
    }

    for key, sheet_name in transport_sheet_map.items():
        df = transport.get(key, pd.DataFrame())
        if df.empty:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(["No data available — check dataset ID in config.py"])
        else:
            _write_df_sheet(wb, sheet_name, df, list(df.columns))

    wb.save(path)
    print(f"  Excel workbook → {path}")
    return zone_df


# ═══════════════════════════════════════════════════════════════════════════════
# 3. data.js — JavaScript data file for the Leaflet website
# ═══════════════════════════════════════════════════════════════════════════════

def _df_to_js_array(df: pd.DataFrame, category: str) -> str:
    """
    Convert a DataFrame to a compact JS array literal string for embedding in JS.
    Only rows with valid coordinates are included.
    If a row has a non-null 'geometry' column (a GeoJSON dict), it is serialized
    as an inline JS object (not a quoted string) in the output.
    """
    parts = []
    for _, row in df.iterrows():
        try:
            lat = float(row.get("latitude"))
            lon = float(row.get("longitude"))
            if math.isnan(lat) or math.isnan(lon):
                continue
        except (TypeError, ValueError):
            continue

        record = {
            "name":     str(row.get("Name", ""))[:120],
            "address":  str(row.get("Address", ""))[:200],
            "type":     str(row.get("Type", ""))[:80],
            "category": str(row.get("category", category)),
            "source":   str(row.get("source", "")),
            "lat":      round(lat, 6),
            "lon":      round(lon, 6),
        }

        geom       = row.get("geometry")      if "geometry"      in row.index else None
        bldg_geom  = row.get("building_geom") if "building_geom" in row.index else None

        has_geom      = geom      is not None and isinstance(geom,      dict)
        has_bldg_geom = bldg_geom is not None and isinstance(bldg_geom, dict)

        if has_geom or has_bldg_geom:
            # Embed geometries as inline JS objects (not quoted strings)
            base = json.dumps(record, ensure_ascii=False, separators=(",", ":"))
            extra = ""
            if has_geom:
                extra += ',"geometry":' + json.dumps(geom, ensure_ascii=False, separators=(",", ":"))
            if has_bldg_geom:
                extra += ',"building_geom":' + json.dumps(bldg_geom, ensure_ascii=False, separators=(",", ":"))
            obj = base[:-1] + extra + "}"
            parts.append(obj)
        else:
            parts.append(json.dumps(record, ensure_ascii=False, separators=(",", ":")))

    return "[" + ",".join(parts) + "]"


def _transport_stops_to_js(df: pd.DataFrame, label: str) -> str:
    """Serialise transport POINT data (stations/stops) to a JS array."""
    records = []
    for _, row in df.iterrows():
        try:
            lat = float(row.get("latitude"))
            lon = float(row.get("longitude"))
            if math.isnan(lat) or math.isnan(lon):
                continue
        except (TypeError, ValueError):
            continue

        records.append({
            "name":  str(row.get("station_name",
                         row.get("stop_name", row.get("name", label))))[:120],
            "label": label,
            "lat":   round(lat, 6),
            "lon":   round(lon, 6),
        })

    return json.dumps(records, ensure_ascii=False, separators=(",", ":"))


def _zones_to_js() -> str:
    """Serialise zone definitions as a JS array of rectangle objects."""
    return json.dumps([
        {
            "key":    k,
            "label":  v["label"],
            "lat_min": v["lat_min"],
            "lat_max": v["lat_max"],
            "lon_min": -87.630,
            "lon_max": -87.525,
            "color":  v["border"],
        }
        for k, v in ZONES.items()
    ], separators=(",", ":"))


def _community_to_geojson(features: list[dict]) -> str:
    """Serialise community area boundary features to a GeoJSON string."""
    if not features:
        return "null"
    return json.dumps({"type": "FeatureCollection", "features": features},
                      separators=(",", ":"))


def _zoning_to_js(zoning: list[dict]) -> str:
    """
    Serialise zoning district records to a compact JSON array string.
    Each element: {zone_class, pd_use, geometry}.
    pd_use is only meaningful for PD zones; empty string otherwise.
    """
    if not zoning:
        return "[]"
    slim = [
        {
            "zone_class": z["zone_class"],
            "pd_use":     z.get("pd_use", ""),
            "geometry":   z["geometry"],
        }
        for z in zoning
    ]
    return json.dumps(slim, separators=(",", ":"))


def export_data_js(assets: dict[str, pd.DataFrame],
                   transport: dict[str, pd.DataFrame],
                   community_features: list[dict],
                   zoning: list[dict] | None = None,
                   vacant_lots: list[dict] | None = None,
                   path: str = OUTPUT_DATA_JS) -> str:
    """
    Generate data.js — a JavaScript file that defines a global
    BRONZEVILLE_DATA constant used by index.html.

    Returns the full JS string (also written to disk).
    """
    # Normalise keys: fetch_assets returns plural (landmarks, restaurants…)
    # but ASSET_CATEGORIES, MARKER_STYLES, and filter checkboxes use singular.
    _PLURAL_TO_SINGULAR = {
        "landmarks": "landmark", "restaurants": "restaurant",
        "businesses": "business", "parks": "park",
    }
    asset_js = {
        _PLURAL_TO_SINGULAR.get(key, key): _df_to_js_array(df, _PLURAL_TO_SINGULAR.get(key, key))
        for key, df in assets.items()
    }

    lines = [
        "// Auto-generated by bronzeville_map.py — do not edit manually.",
        "// Re-run main.py to refresh this file.",
        "",
        "const BRONZEVILLE_DATA = {",
    ]

    # Asset categories
    lines.append("  assets: {")
    for key, js_arr in asset_js.items():
        lines.append(f"    {key}: {js_arr},")
    lines.append("  },")

    # Transport stops (point data only — lines are complex geometries)
    lines.append("  transport: {")
    lines.append(f"    cta_rail: {_transport_stops_to_js(transport.get('cta_rail_stations', pd.DataFrame()), 'CTA Rail Station')},")
    lines.append(f"    cta_bus:  {_transport_stops_to_js(transport.get('cta_bus_stops', pd.DataFrame()), 'CTA Bus Stop')},")
    lines.append(f"    metra:    {_transport_stops_to_js(transport.get('metra_stations', pd.DataFrame()), 'Metra Station')},")
    lines.append("  },")

    # Zones and community boundary
    lines.append(f"  zones: {_zones_to_js()},")
    lines.append(f"  communityBoundary: {_community_to_geojson(community_features)},")

    # Zoning districts (SimCity layer)
    zoning_js = _zoning_to_js(zoning or [])
    lines.append(f"  zoning: {zoning_js},")

    # Vacant lots overlay
    import json as _json
    vacant_js = _json.dumps(vacant_lots or [], separators=(",", ":"))
    lines.append(f"  vacantLots: {vacant_js},")

    lines.append("};")

    js_content = "\n".join(lines)

    with open(path, "w", encoding="utf-8") as f:
        f.write(js_content)

    print(f"  data.js       → {path}")
    return js_content


# ═══════════════════════════════════════════════════════════════════════════════
# 4. index.html — Standalone Leaflet.js Website
# ═══════════════════════════════════════════════════════════════════════════════

def _build_filter_html() -> str:
    """Generate the display-mode selector and filter-panel checkboxes."""
    transport_layers = [
        ("cta_rail",   "cta_rail",   "CTA Rail Stations"),
        ("cta_bus",    "cta_bus",    "CTA Bus Stops"),
        ("metra",      "metra",      "Metra Stations"),
    ]

    # Display mode selector — rendered above the category checkboxes
    mode_html = (
        '<div class="mode-selector">'
        '<div class="mode-label">Basemap</div>'
        '<div class="mode-options">'
        '<label><input type="radio" name="basemap" value="street" checked> Street</label>'
        '<label><input type="radio" name="basemap" value="satellite"> Satellite</label>'
        '</div>'
        '</div>'
    )

    html = mode_html + '<div class="filter-section"><b>Community Assets</b>'
    for cat in ASSET_CATEGORIES:
        style  = MARKER_STYLES.get(cat, MARKER_STYLES["landmark"])
        label  = style["label"]
        hex_c  = style["hex"]
        html += (
            f'<label class="filter-label">'
            f'<input type="checkbox" class="layer-toggle" data-layer="{cat}" checked> '
            f'<span class="dot" style="background:{hex_c}"></span> {label}'
            f'</label>'
        )
    # Vacant lots / storefronts / abandoned structures — moved from Map Overlays
    html += (
        '<label class="filter-label">'
        '<input type="checkbox" class="layer-toggle" data-layer="vacant" checked> '
        '<span class="dot" style="background:#795548"></span> Vacant Lots &amp; Buildings'
        '</label>'
    )

    html += '</div><div class="filter-section"><b>Transportation</b>'
    for key, cat, label in transport_layers:
        style = MARKER_STYLES.get(cat, MARKER_STYLES["cta_rail"])
        hex_c = style["hex"]
        html += (
            f'<label class="filter-label">'
            f'<input type="checkbox" class="layer-toggle" data-layer="{key}" checked> '
            f'<span class="dot" style="background:{hex_c}"></span> {label}'
            f'</label>'
        )
    html += '</div>'

    # SimCity zoning overlay toggle
    html += (
        '<div class="filter-section simcity-section">'
        '<b>Map Overlays</b>'
        '<label class="filter-label simcity-label">'
        '<input type="checkbox" id="simcity-toggle"> '
        '<span class="simcity-swatch"></span> '
        'SimCity Zoning'
        '</label>'
        '<div id="zoning-legend" style="display:none;margin-top:6px;font-size:11px">'
        '<div><span class="dot" style="background:#66BB6A"></span> Residential</div>'
        '<div><span class="dot" style="background:#9C6FDE"></span> Mixed-Use (B-class)</div>'
        '<div><span class="dot" style="background:#42A5F5"></span> Commercial</div>'
        '<div><span class="dot" style="background:#FFCA28"></span> Industrial</div>'
        '<div><span class="dot" style="background:#2E7D32"></span> Parks / Open Space</div>'
        '<div><span class="dot" style="background:#78909C"></span> Transportation</div>'
        '<div style="margin-top:4px;color:#888;font-style:italic">Planned Developments</div>'
        '<div><span class="dot" style="background:#66BB6A"></span> PD — Residential</div>'
        '<div><span class="dot" style="background:#9C6FDE"></span> PD — Mixed-Use</div>'
        '<div><span class="dot" style="background:#42A5F5"></span> PD — Commercial</div>'
        '<div><span class="dot" style="background:#FFCA28"></span> PD — Industrial</div>'
        '<div><span class="dot" style="background:#2E7D32"></span> PD — Park</div>'
        '<div><span class="dot" style="background:#BDBDBD"></span> PD — Unclassified</div>'
        '</div>'
        '</label>'
        '</div>'
    )
    return html


def _build_legend_html() -> str:
    """Generate the map legend HTML."""
    items = []
    for cat in ASSET_CATEGORIES:
        s = MARKER_STYLES[cat]
        items.append(
            f'<div class="legend-item">'
            f'<span class="legend-dot" style="background:{s["hex"]}"></span>'
            f' {s["label"]}'
            f'</div>'
        )
    for cat in ("cta_rail", "cta_bus", "metra"):
        s = MARKER_STYLES[cat]
        items.append(
            f'<div class="legend-item">'
            f'<span class="legend-dot" style="background:{s["hex"]}"></span>'
            f' {s["label"]}'
            f'</div>'
        )
    return "\n".join(items)


_HTML_TEMPLATE = """\
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>Bronzeville Community Asset Map</title>

  <!-- Leaflet CSS -->
  <link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css" />
  <!-- Leaflet MarkerCluster CSS -->
  <link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.css" />
  <link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.Default.css" />
  <!-- Font Awesome (icons) -->
  <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css" />

  <style>
    /* ── Reset ───────────────────────────────────────────── */
    *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
    body { font-family: 'Segoe UI', Arial, sans-serif; display: flex;
           flex-direction: column; height: 100vh; overflow: hidden; }

    /* ── Header ─────────────────────────────────────────── */
    #header {
      background: #1a1a2e; color: #fff;
      padding: 10px 18px; display: flex;
      align-items: center; justify-content: space-between;
      flex-shrink: 0; z-index: 1001;
    }
    #header h1 { font-size: 1.15rem; font-weight: 700; letter-spacing: .03em; }
    #header .subtitle { font-size: .8rem; color: #aab; margin-top: 2px; }

    /* ── Search box ─────────────────────────────────────── */
    #search-wrap { position: relative; }
    #search-box {
      padding: 6px 12px; border-radius: 20px;
      border: none; font-size: .9rem; width: 220px;
      background: rgba(255,255,255,.15); color: #fff;
      outline: none;
    }
    #search-box::placeholder { color: #ccc; }
    #search-box:focus { background: rgba(255,255,255,.25); }
    #search-clear {
      position: absolute; right: 10px; top: 50%;
      transform: translateY(-50%); cursor: pointer;
      color: #ccc; font-size: .8rem; display: none;
    }

    /* ── Main layout ────────────────────────────────────── */
    #main { display: flex; flex: 1; overflow: hidden; }

    /* ── Sidebar ────────────────────────────────────────── */
    #sidebar {
      width: 230px; flex-shrink: 0;
      background: #f8f9fa; border-right: 1px solid #dde;
      overflow-y: auto; padding: 14px 12px;
      display: flex; flex-direction: column; gap: 12px;
    }
    .filter-section { display: flex; flex-direction: column; gap: 5px; }
    .filter-section b { font-size: .8rem; text-transform: uppercase;
                        color: #555; letter-spacing: .06em; margin-bottom: 2px; }
    .filter-label { display: flex; align-items: center; gap: 7px;
                    font-size: .88rem; cursor: pointer; padding: 3px 0; }
    .filter-label:hover { color: #1a1a2e; }
    .dot { display: inline-block; width: 12px; height: 12px;
           border-radius: 50%; flex-shrink: 0; }

    /* ── Display mode selector ──────────────────────────── */
    .mode-selector {
      background: #eef0f8; border-radius: 7px;
      padding: 9px 11px; border: 1px solid #d0d4e8;
    }
    .mode-label {
      font-size: .8rem; text-transform: uppercase;
      color: #555; letter-spacing: .06em;
      font-weight: 700; margin-bottom: 6px;
    }
    .mode-options { display: flex; flex-direction: column; gap: 4px; }
    .mode-options label {
      font-size: .88rem; cursor: pointer;
      display: flex; align-items: center; gap: 6px;
    }
    .mode-options label:hover { color: #1a1a2e; }
    .mode-options input[type="radio"] { accent-color: #1a1a2e; }

    /* SimCity toggle swatch */
    .simcity-section { border-top: 1px solid #e0e0e8; padding-top: 8px; margin-top: 4px; }
    .simcity-swatch {
      display: inline-block; width: 22px; height: 12px;
      border-radius: 3px; flex-shrink: 0;
      background: linear-gradient(90deg, #66BB6A 0%, #42A5F5 40%, #FFCA28 80%, #2E7D32 100%);
    }
    .simcity-label { font-weight: 600; }
    #zoning-legend .dot { width: 10px; height: 10px; margin-right: 4px; }
    #zoning-legend div { display: flex; align-items: center; padding: 1px 0; }

    /* Restrict-to-boundary toggle */
    #restrict-btn {
      margin-top: 4px; padding: 7px 10px;
      background: #1a1a2e; color: #fff;
      border: none; border-radius: 6px; cursor: pointer;
      font-size: .85rem; text-align: center;
    }
    #restrict-btn:hover { background: #2c2c5e; }

    /* Count badges */
    #counts { font-size: .8rem; color: #666; }
    #counts table { width: 100%; border-collapse: collapse; }
    #counts td { padding: 2px 4px; }
    #counts td:last-child { text-align: right; font-weight: 600; color: #333; }

    /* ── Map ───────────────────────────────────────────── */
    #map { flex: 1; }

    /* ── Legend (bottom-right) ─────────────────────────── */
    #legend {
      position: absolute; bottom: 30px; right: 10px;
      background: rgba(255,255,255,.93);
      border-radius: 8px; padding: 10px 13px;
      box-shadow: 0 2px 8px rgba(0,0,0,.2);
      z-index: 1000; font-size: .82rem; line-height: 1.6;
      max-height: 55vh; overflow-y: auto;
    }
    #legend b { display: block; margin-bottom: 4px; font-size: .85rem;
                text-transform: uppercase; letter-spacing: .05em; color: #333; }
    .legend-item { display: flex; align-items: center; gap: 7px; }
    .legend-dot { width: 11px; height: 11px; border-radius: 50%;
                  flex-shrink: 0; display: inline-block; }

    /* ── Popup ──────────────────────────────────────────── */
    .leaflet-popup-content b { font-size: 1em; }
    .leaflet-popup-content .meta { color: #666; font-size: .85em; }
    .leaflet-popup-content hr { border-color: #eee; margin: 5px 0; }

    /* ── Mobile drawer button (hidden on desktop) ──────── */
    #drawer-btn {
      display: none;
      background: none; border: none; cursor: pointer;
      color: #fff; font-size: 1.4rem; padding: 4px 8px;
      line-height: 1; margin-right: 8px;
    }

    /* ── Overlay backdrop ───────────────────────────────── */
    #drawer-overlay {
      display: none; position: fixed; inset: 0;
      background: rgba(0,0,0,.45); z-index: 1099;
    }
    #drawer-overlay.open { display: block; }

    /* ── Responsive ────────────────────────────────────── */
    @media (max-width: 640px) {
      #drawer-btn { display: block; }
      #legend     { display: none; }

      /* Sidebar becomes a slide-in drawer */
      #sidebar {
        position: fixed; top: 0; left: 0; height: 100%;
        width: 270px; z-index: 1100;
        transform: translateX(-100%);
        transition: transform .25s ease;
        box-shadow: 3px 0 12px rgba(0,0,0,.25);
      }
      #sidebar.open { transform: translateX(0); }
    }
  </style>
</head>

<body>

<!-- ── Header ──────────────────────────────────────────────── -->
<div id="header">
  <div style="display:flex;align-items:center">
    <button id="drawer-btn" aria-label="Open filters">&#9776;</button>
    <div>
      <h1>Bronzeville Community Asset Map</h1>
      <div class="subtitle">Chicago, Illinois — Community Asset Inventory</div>
    </div>
  </div>
  <div id="search-wrap">
    <input id="search-box" type="text" placeholder="Search assets…" autocomplete="off" />
    <span id="search-clear" title="Clear search">✕</span>
  </div>
</div>

<!-- ── Main Layout ─────────────────────────────────────────── -->
<div id="main">

  <!-- Sidebar filter panel -->
  <div id="sidebar">
    __FILTER_HTML__

    <button id="restrict-btn">📍 Restrict to Bronzeville</button>

    <hr style="border-color:#ddd;">
    <div id="counts">
      <b style="font-size:.8rem;text-transform:uppercase;color:#555;letter-spacing:.06em;">
        Visible Assets
      </b>
      <table id="count-table"></table>
    </div>
  </div>

  <!-- Map container -->
  <div id="map"></div>

</div>

<!-- ── Mobile drawer overlay ────────────────────────────────── -->
<div id="drawer-overlay"></div>

<!-- Legend (overlaid on map) -->
<div id="legend">
  <b>Legend</b>
  __LEGEND_HTML__
</div>

<!-- ── Scripts ────────────────────────────────────────────── -->
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script src="https://unpkg.com/leaflet.markercluster@1.5.3/dist/leaflet.markercluster.js"></script>

<!-- Inline data (auto-generated — re-run main.py to refresh) -->
<script>
__INLINE_DATA__
</script>

<script>
// ═══════════════════════════════════════════════════════════════
//  MAP INITIALISATION
// ═══════════════════════════════════════════════════════════════

const MAP_CENTER = [__LAT__, __LON__];
const MAP_ZOOM   = 14;

const map = L.map('map', { center: MAP_CENTER, zoom: MAP_ZOOM });

// Base tile layers
const baseLayers = {
  street: L.tileLayer('https://{s}.basemaps.cartocdn.com/light_all/{z}/{x}/{y}{r}.png', {
    attribution: '© <a href="https://www.openstreetmap.org/">OpenStreetMap</a> contributors © <a href="https://carto.com/">CARTO</a>',
    subdomains: 'abcd',
    maxZoom: 19,
  }),
  satellite: L.tileLayer('https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}', {
    attribution: '© <a href="https://www.esri.com/">Esri</a>, Maxar, Earthstar Geographics',
    maxZoom: 19,
  }),
};
baseLayers.street.addTo(map);
let activeBaseLayer = baseLayers.street;

// ── Community boundary polygon ──────────────────────────────
let communityLayer = null;
let communityBounds = null;

if (BRONZEVILLE_DATA.communityBoundary) {
  communityLayer = L.geoJSON(BRONZEVILLE_DATA.communityBoundary, {
    style: {
      color: '#333', weight: 3, dashArray: '6 4',
      fillColor: '#000', fillOpacity: 0.04,
    },
    onEachFeature(feature, layer) {
      if (feature.properties && feature.properties.community) {
        layer.bindTooltip(feature.properties.community, { sticky: true });
      }
    },
  }).addTo(map);
  communityBounds = communityLayer.getBounds();
}

// Zone rectangles removed — boundary polygon provides the geographic context.

// ═══════════════════════════════════════════════════════════════
//  MARKER FACTORY
// ═══════════════════════════════════════════════════════════════

// Colour map keyed by category (matches config.py MARKER_STYLES)
const COLOURS = {
  landmark:   '#1A5276',
  restaurant: '#C0392B',
  business:   '#1E8449',
  education:  '#6C3483',
  worship:    '#CA6F1E',
  park:       '#145A32',
  health:     '#922B21',
  cta_rail:   '#455A64',
  cta_bus:    '#455A64',
  metra:      '#455A64',
};

// SimCity asset colours — used when SimCity zoning mode is active
const SIMCITY_COLOURS = {
  landmark:   '#F9A825',   // Gold
  restaurant: '#FF7043',   // Coral
  business:   '#00897B',   // Teal
  education:  '#7B1FA2',   // Purple
  worship:    '#E91E63',   // Pink
  park:       '#1B5E20',   // Dark Green
  health:     '#D32F2F',   // Red
  cta_rail:   '#455A64',   // Transit slate
  cta_bus:    '#455A64',   // Transit slate
  metra:      '#455A64',   // Transit slate
};

// Active colour map — switched by SimCity toggle
let currentColorMap = COLOURS;

// ── SimCity zoning helpers ──────────────────────────────────
const PD_USE_COLORS = {
  residential:   '#66BB6A',   // Green
  commercial:    '#42A5F5',   // Blue
  industrial:    '#FFCA28',   // Yellow
  park:          '#2E7D32',   // Dark Green
  mixed:         '#AB47BC',   // Purple (mix of residential + commercial)
  transportation:'#78909C',   // Blue-gray
  unknown:       '#BDBDBD',   // Gray fallback
};

function getZoneColor(z) {
  const zc = (z.zone_class || '').toUpperCase();
  // Residential (single, two-flat, multi-family)
  if (/^R[STM]|^DR|^RA/.test(zc))     return '#66BB6A';
  // Business class = mixed-use (commercial ground floor + residential above)
  if (/^B[123]/.test(zc))              return '#9C6FDE';
  // Commercial (more purely commercial)
  if (/^C[123]/.test(zc))              return '#42A5F5';
  // Industrial / Manufacturing
  if (/^M[123]|^PMD/.test(zc))         return '#FFCA28';
  // Parks & Open Space
  if (/^POS|^POC/.test(zc))            return '#2E7D32';
  // Planned Development — use derived pd_use
  if (/^PD/.test(zc))                  return PD_USE_COLORS[z.pd_use] || '#BDBDBD';
  // Transportation
  if (zc === 'T')                       return '#78909C';
  return '#E0E0E0';
}

let simCityMode = false;
let zoningLayer  = null;

// Inject stripe SVG pattern into Leaflet's overlay pane SVG.
// Called before buildZoningLayer so the pattern exists when polygons render.
function injectMixedZonePattern() {
  const tryInject = () => {
    const svg = document.querySelector('.leaflet-overlay-pane svg');
    if (!svg) { setTimeout(tryInject, 60); return; }
    if (svg.querySelector('#stripe-mixed')) return;  // already injected

    const NS   = 'http://www.w3.org/2000/svg';
    let defs   = svg.querySelector('defs');
    if (!defs) {
      defs = document.createElementNS(NS, 'defs');
      svg.insertBefore(defs, svg.firstChild);
    }

    const pat = document.createElementNS(NS, 'pattern');
    pat.setAttribute('id',               'stripe-mixed');
    pat.setAttribute('x',                '0');
    pat.setAttribute('y',                '0');
    pat.setAttribute('width',            '10');
    pat.setAttribute('height',           '10');
    pat.setAttribute('patternUnits',     'userSpaceOnUse');
    pat.setAttribute('patternTransform', 'rotate(45)');

    const r1 = document.createElementNS(NS, 'rect');
    r1.setAttribute('width',  '5');
    r1.setAttribute('height', '10');
    r1.setAttribute('fill',   '#66BB6A');

    const r2 = document.createElementNS(NS, 'rect');
    r2.setAttribute('x',      '5');
    r2.setAttribute('width',  '5');
    r2.setAttribute('height', '10');
    r2.setAttribute('fill',   '#42A5F5');

    pat.appendChild(r1);
    pat.appendChild(r2);
    defs.appendChild(pat);
  };
  tryInject();
}

function buildZoningLayer() {
  const zData = BRONZEVILLE_DATA.zoning || [];
  if (!zData.length) return null;
  return L.geoJSON({
    type: 'FeatureCollection',
    features: zData.map(z => ({
      type:       'Feature',
      geometry:    z.geometry,
      properties: { zone_class: z.zone_class, pd_use: z.pd_use || '' },
    }))
  }, {
    style: f => {
      return {
        fillColor:   getZoneColor(f.properties),
        fillOpacity: 0.50,
        color:       '#ffffff',
        weight:      0.5,
        opacity:     0.35,
      };
    },
    onEachFeature: (feature, layer) => {
      const zc  = feature.properties.zone_class;
      const tip = feature.properties.pd_use
        ? `<b>${zc}</b> <span style="color:#888">(${feature.properties.pd_use})</span>`
        : `<b>${zc}</b>`;
      layer.bindTooltip(tip, { sticky: true, direction: 'top' });
    },
  });
}

function setSimCityMode(enabled) {
  simCityMode      = enabled;
  currentColorMap  = enabled ? SIMCITY_COLOURS : COLOURS;

  // Show / hide zoning polygon layer
  if (enabled) {
    if (!zoningLayer) zoningLayer = buildZoningLayer();
    if (zoningLayer) { zoningLayer.addTo(map); zoningLayer.bringToBack(); }
  } else {
    if (zoningLayer) zoningLayer.remove();
  }

  // Rebuild asset layers with new colour palette
  rebuildAssetLayers();

  // Show / hide zoning legend
  const leg = document.getElementById('zoning-legend');
  if (leg) leg.style.display = enabled ? 'block' : 'none';
}

// Font-Awesome icon class per category
const ICONS = {
  landmark:   'fa-star',
  restaurant: 'fa-utensils',
  business:   'fa-briefcase',
  education:  'fa-graduation-cap',
  worship:    'fa-church',
  park:       'fa-tree',
  health:     'fa-plus-square',
  cta_rail:   'fa-train',
  cta_bus:    'fa-bus',
  metra:      'fa-train',
};

function makeIcon(category) {
  const color = COLOURS[category] || '#555';
  const icon  = ICONS[category]   || 'fa-map-marker';
  return L.divIcon({
    className: '',
    iconSize:  [28, 28],
    iconAnchor:[14, 14],
    popupAnchor:[0, -14],
    html: `<div style="
      width:28px;height:28px;border-radius:50%;
      background:${color};display:flex;align-items:center;
      justify-content:center;box-shadow:0 1px 4px rgba(0,0,0,.4);
      border:2px solid #fff;">
      <i class="fa ${icon}" style="color:#fff;font-size:12px;"></i>
    </div>`,
  });
}

function makePopup(d) {
  return `
    <div style="min-width:180px;max-width:250px;font-family:Arial,sans-serif">
      <b>${d.name || '—'}</b><br>
      <span class="meta">${d.address || 'Address unavailable'}</span>
      <hr>
      Category: <b>${(d.category || '').replace(/_/g,' ')}</b><br>
      Type: ${d.type || '—'}<br>
      <small style="color:#aaa">Source: ${d.source || ''}</small>
    </div>`;
}

// ═══════════════════════════════════════════════════════════════
//  LAYER GROUPS
// ═══════════════════════════════════════════════════════════════

// One MarkerClusterGroup per asset category + one per transport type
const layerGroups = {};

// ── Point-in-polygon helper (ray casting) ──────────────────
// Used as a client-side guard to ensure no markers appear outside
// the community boundary, regardless of data source.
function pointInGeoJSON(lon, lat, geojson) {
  if (!geojson) return true;  // no boundary — show everything
  const features = geojson.features || [];
  return features.some(f => {
    const g = f.geometry;
    if (!g) return false;
    const polys = g.type === 'Polygon'
      ? [g.coordinates]
      : g.type === 'MultiPolygon'
        ? g.coordinates
        : [];
    return polys.some(poly => raycast(lon, lat, poly[0]));
  });
}

function raycast(x, y, ring) {
  let inside = false;
  for (let i = 0, j = ring.length - 1; i < ring.length; j = i++) {
    const xi = ring[i][0], yi = ring[i][1];
    const xj = ring[j][0], yj = ring[j][1];
    if (((yi > y) !== (yj > y)) && (x < (xj - xi) * (y - yi) / (yj - yi) + xi))
      inside = !inside;
  }
  return inside;
}

// Current display mode — 'parcels' | 'buildings' | 'pins'
let currentMode = 'pins';

function _renderAsset(d, color, group) {
  // Choose the geometry source based on currentMode, with fallbacks
  let geomSource = null;
  if (currentMode === 'parcels') {
    geomSource = d.geometry || null;
  } else if (currentMode === 'buildings') {
    geomSource = d.building_geom || null;
  }
  // 'pins' mode always uses circleMarker; polygon modes fall back to circleMarker if no geom

  if (geomSource) {
    L.geoJSON(geomSource, {
      style: {
        color:       color,
        fillColor:   color,
        fillOpacity: currentMode === 'buildings' ? 0.45 : 0.35,
        weight:      2,
      },
    })
    .bindPopup(makePopup(d))
    .bindTooltip(d.name, { direction: 'top', sticky: true })
    .addTo(group);
  } else {
    // Pin fallback (also used for 'pins' mode)
    L.circleMarker([d.lat, d.lon], {
      radius:      currentMode === 'pins' ? 6 : 5,
      color:       color,
      fillColor:   color,
      fillOpacity: currentMode === 'pins' ? 0.85 : 0.7,
      weight:      1.5,
    })
    .bindPopup(makePopup(d))
    .bindTooltip(d.name, { direction: 'top' })
    .addTo(group);
  }
}

function buildAssetLayers() {
  const boundary = BRONZEVILLE_DATA.communityBoundary;
  const categories = Object.keys(BRONZEVILLE_DATA.assets);
  categories.forEach(cat => {
    const color = currentColorMap[cat] || COLOURS[cat] || '#555';
    // Use plain layerGroup — polygons don't cluster well
    const group = L.layerGroup();
    (BRONZEVILLE_DATA.assets[cat] || []).forEach(d => {
      if (!d.lat || !d.lon) return;
      if (!pointInGeoJSON(d.lon, d.lat, boundary)) return;
      _renderAsset(d, color, group);
    });
    layerGroups[cat] = group;
    group.addTo(map);
  });
}

function rebuildAssetLayers() {
  // Remove and recreate all asset layers, preserving visibility state
  const categories = Object.keys(BRONZEVILLE_DATA.assets);
  categories.forEach(cat => {
    const existing = layerGroups[cat];
    if (existing) {
      if (map.hasLayer(existing)) existing.remove();
      delete layerGroups[cat];
    }
  });
  buildAssetLayers();
  // Restore visibility state for each category
  categories.forEach(cat => {
    const visible = layerVisible[cat];
    // layerVisible[cat] may be undefined if never toggled — default to true (checked)
    if (visible === false) {
      const group = layerGroups[cat];
      if (group && map.hasLayer(group)) group.remove();
    }
  });
  updateCounts();
}

function buildTransportLayers() {
  const tData = BRONZEVILLE_DATA.transport;

  // CTA bus stops — plain small dots, no clustering
  const busGroup = L.layerGroup();
  (tData['cta_bus'] || []).forEach(d => {
    if (!d.lat || !d.lon) return;
    L.circleMarker([d.lat, d.lon], {
      radius: 3, color: COLOURS['cta_bus'], fillColor: COLOURS['cta_bus'],
      fillOpacity: 0.75, weight: 1,
    })
    .bindPopup(`<b>${d.name}</b><br><i>${d.label}</i>`)
    .addTo(busGroup);
  });
  layerGroups['cta_bus'] = busGroup;
  busGroup.addTo(map);

  // CTA rail stations and Metra — plain markers, no clustering
  ['cta_rail', 'metra'].forEach(key => {
    const group = L.layerGroup();
    (tData[key] || []).forEach(d => {
      if (!d.lat || !d.lon) return;
      L.circleMarker([d.lat, d.lon], {
        radius: 6, color: COLOURS[key], fillColor: COLOURS[key],
        fillOpacity: 0.85, weight: 1.5,
      })
      .bindPopup(`<b>${d.name}</b><br><i>${d.label}</i>`)
      .addTo(group);
    });
    layerGroups[key] = group;
    group.addTo(map);
  });
}

// ═══════════════════════════════════════════════════════════════
//  FILTER PANEL
// ═══════════════════════════════════════════════════════════════

// Explicit visibility state map — more reliable than removeLayer with cluster plugins
const layerVisible = {};

function setLayerVisible(key, visible) {
  layerVisible[key] = visible;
  const group = layerGroups[key];
  if (!group) return;
  if (visible) {
    if (!map.hasLayer(group)) group.addTo(map);
  } else {
    if (map.hasLayer(group)) group.remove();
  }
}

buildAssetLayers();
buildTransportLayers();
buildVacantLayer();

document.querySelectorAll('.layer-toggle').forEach(cb => {
  const key = cb.dataset.layer;
  // Initialise state to match the checkbox's starting value
  layerVisible[key] = cb.checked;

  cb.addEventListener('change', function () {
    setLayerVisible(this.dataset.layer, this.checked);
    updateCounts();
  });
});

// ── Basemap toggle ──────────────────────────────────────────
document.querySelectorAll('input[name="basemap"]').forEach(radio => {
  radio.addEventListener('change', function () {
    if (!this.checked) return;
    map.removeLayer(activeBaseLayer);
    activeBaseLayer = baseLayers[this.value];
    activeBaseLayer.addTo(map);
    activeBaseLayer.bringToBack();
  });
});

// ── SimCity zoning toggle ───────────────────────────────────
const simCityToggle = document.getElementById('simcity-toggle');
if (simCityToggle) {
  simCityToggle.addEventListener('change', function () {
    setSimCityMode(this.checked);
  });
}

// ── Vacant / Abandoned layer ────────────────────────────────
// Three sub-types with distinct colours:
//   vacant_land        → sandy brown  #D4956A  (bare lots, no structure)
//   vacant_storefront  → orange       #E67E22  (empty commercial buildings)
//   abandoned_structure→ brick red    #C0392B  (311-reported derelict buildings)

function buildVacantLayer() {
  const lots = (BRONZEVILLE_DATA.vacantLots || []);
  if (!lots.length) return;
  const group = L.layerGroup();
  const STYLES = {
    vacant_land:       { radius: 4, color: '#A0522D', fillColor: '#D4956A' },
    vacant_storefront: { radius: 5, color: '#C0640A', fillColor: '#E67E22' },
  };
  const DEFAULT_STYLE = { radius: 4, color: '#795548', fillColor: '#A1887F' };

  lots.forEach(function(lot) {
    const style     = STYLES[lot.type] || DEFAULT_STYLE;
    const address   = lot.address    || 'Address unknown';
    const typeLabel = lot.status     || 'Vacant';
    const classCode = lot.class_code || '';
    const dot = L.circleMarker([lot.lat, lot.lon], {
      radius:      style.radius,
      color:       style.color,
      fillColor:   style.fillColor,
      fillOpacity: 0.85,
      weight:      1,
    });
    dot.bindPopup(
      '<div style="min-width:160px">'
      + '<b style="font-size:13px">' + address + '</b><hr style="margin:4px 0">'
      + '<b>' + typeLabel + '</b>'
      + (classCode ? '<br><small style="color:#888">Cook County Class ' + classCode + '</small>' : '')
      + '</div>'
    );
    dot.addTo(group);
  });
  layerGroups['vacant'] = group;
  group.addTo(map);
}

// ── Restrict-to-boundary button ────────────────────────────
let restricted = false;
document.getElementById('restrict-btn').addEventListener('click', function () {
  restricted = !restricted;
  if (restricted && communityBounds) {
    map.fitBounds(communityBounds, { padding: [20, 20] });
    this.textContent = '🗺️ Show Full View';
  } else {
    map.setView(MAP_CENTER, MAP_ZOOM);
    this.textContent = '📍 Restrict to Bronzeville';
  }
});

// ═══════════════════════════════════════════════════════════════
//  SEARCH
// ═══════════════════════════════════════════════════════════════

let searchMarkers = [];

function clearSearch() {
  searchMarkers.forEach(m => map.removeLayer(m));
  searchMarkers = [];
  document.getElementById('search-clear').style.display = 'none';
}

function doSearch(query) {
  clearSearch();
  if (!query.trim()) return;

  const q = query.toLowerCase();
  let found = 0;
  const bounds = [];

  // Search all asset categories
  Object.entries(BRONZEVILLE_DATA.assets).forEach(([cat, items]) => {
    (items || []).forEach(d => {
      if (!d.lat || !d.lon) return;
      if ((d.name || '').toLowerCase().includes(q) ||
          (d.address || '').toLowerCase().includes(q) ||
          (d.type || '').toLowerCase().includes(q)) {
        const m = L.marker([d.lat, d.lon], { icon: makeIcon(cat), zIndexOffset: 1000 })
          .bindPopup(makePopup(d))
          .addTo(map)
          .openPopup();
        searchMarkers.push(m);
        bounds.push([d.lat, d.lon]);
        found++;
      }
    });
  });

  // Also search transport stops
  Object.entries(BRONZEVILLE_DATA.transport).forEach(([key, items]) => {
    (items || []).forEach(d => {
      if (!d.lat || !d.lon) return;
      if ((d.name || '').toLowerCase().includes(q)) {
        const m = L.circleMarker([d.lat, d.lon], {
          radius: 10, color: COLOURS[key], fillColor: COLOURS[key],
          fillOpacity: 0.9, weight: 2, zIndexOffset: 1000,
        })
        .bindPopup(`<b>${d.name}</b><br><i>${d.label}</i>`)
        .addTo(map);
        searchMarkers.push(m);
        bounds.push([d.lat, d.lon]);
        found++;
      }
    });
  });

  if (bounds.length > 0) {
    if (bounds.length === 1) {
      map.setView(bounds[0], 16);
    } else {
      map.fitBounds(bounds, { padding: [30, 30] });
    }
  }

  document.getElementById('search-clear').style.display = found ? 'inline' : 'none';
}

const searchBox   = document.getElementById('search-box');
const searchClear = document.getElementById('search-clear');

let searchTimeout;
searchBox.addEventListener('input', function () {
  clearTimeout(searchTimeout);
  searchTimeout = setTimeout(() => doSearch(this.value), 300);
});

searchBox.addEventListener('keydown', function (e) {
  if (e.key === 'Escape') { this.value = ''; clearSearch(); }
});

searchClear.addEventListener('click', function () {
  searchBox.value = '';
  clearSearch();
});

// ═══════════════════════════════════════════════════════════════
//  ASSET COUNT TABLE
// ═══════════════════════════════════════════════════════════════

function updateCounts() {
  const tbody = document.getElementById('count-table');
  tbody.innerHTML = '';

  const allCats = [
    ...Object.keys(BRONZEVILLE_DATA.assets),
    'cta_rail', 'cta_bus', 'metra',
  ];

  allCats.forEach(cat => {
    const cb = document.querySelector(`.layer-toggle[data-layer="${cat}"]`);
    if (!cb || !cb.checked) return;

    let count = 0;
    if (BRONZEVILLE_DATA.assets[cat]) {
      count = BRONZEVILLE_DATA.assets[cat].length;
    } else if (BRONZEVILLE_DATA.transport[cat]) {
      count = BRONZEVILLE_DATA.transport[cat].length;
    }

    const label = cat.replace(/_/g, ' ').replace(/\\b\\w/g, c => c.toUpperCase());
    const color = COLOURS[cat] || '#555';
    tbody.innerHTML +=
      `<tr>
        <td><span style="color:${color}">●</span> ${label}</td>
        <td>${count}</td>
      </tr>`;
  });
}

updateCounts();

// ═══════════════════════════════════════════════════════════════
//  MOBILE DRAWER
// ═══════════════════════════════════════════════════════════════

const drawerBtn     = document.getElementById('drawer-btn');
const sidebar       = document.getElementById('sidebar');
const drawerOverlay = document.getElementById('drawer-overlay');

function openDrawer() {
  sidebar.classList.add('open');
  drawerOverlay.classList.add('open');
  document.body.style.overflow = 'hidden';
}

function closeDrawer() {
  sidebar.classList.remove('open');
  drawerOverlay.classList.remove('open');
  document.body.style.overflow = '';
}

if (drawerBtn) {
  drawerBtn.addEventListener('click', function () {
    sidebar.classList.contains('open') ? closeDrawer() : openDrawer();
  });
}

if (drawerOverlay) {
  drawerOverlay.addEventListener('click', closeDrawer);
}

</script>
</body>
</html>
"""


def export_index_html(assets: dict[str, pd.DataFrame],
                      transport: dict[str, pd.DataFrame],
                      community_features: list[dict],
                      zoning: list[dict] | None = None,
                      vacant_lots: list[dict] | None = None,
                      path: str = OUTPUT_INDEX_HTML) -> None:
    """
    Generate a fully self-contained index.html with the asset data embedded
    inline. No server required — open directly in a browser.

    Also writes data.js for use with a local HTTP server.
    """
    # Generate the data.js content and also store it inline
    js_data_content = export_data_js(assets, transport, community_features, zoning=zoning, vacant_lots=vacant_lots)

    filter_html = _build_filter_html()
    legend_html = _build_legend_html()

    lat, lon = BRONZEVILLE_CENTER

    html = (
        _HTML_TEMPLATE
        .replace("__FILTER_HTML__",  filter_html)
        .replace("__LEGEND_HTML__",  legend_html)
        .replace("__INLINE_DATA__",  js_data_content)
        .replace("__LAT__",          str(lat))
        .replace("__LON__",          str(lon))
    )

    with open(path, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"  index.html    → {path}")


# ═══════════════════════════════════════════════════════════════════════════════
# 5. HTML Summary Tables
# ═══════════════════════════════════════════════════════════════════════════════

def export_summary_html(assets: dict[str, pd.DataFrame],
                        path: str = OUTPUT_SUMMARY) -> None:
    """
    Write a formatted HTML file with a summary table for each asset category.
    """
    cols = ["Name", "Address", "Type", "latitude", "longitude", "source"]

    css = """
    body { font-family: Arial, sans-serif; margin: 2em; color: #222; }
    h1   { color: #1a1a2e; }
    h2   { color: #2c3e50; margin-top: 2em; border-bottom: 2px solid #dde; }
    table { border-collapse: collapse; width: 100%; margin: .8em 0 2em; }
    th   { background: #1a1a2e; color: #fff; padding: 7px 10px; text-align: left; }
    td   { padding: 5px 10px; border-bottom: 1px solid #eee; font-size: .9em; }
    tr:nth-child(even) td { background: #f6f7fb; }
    """

    sections = []
    for key, label in [
        ("landmarks",   "Landmarks"),
        ("restaurants", "Restaurants"),
        ("businesses",  "Businesses"),
        ("education",   "Education & Culture"),
        ("worship",     "Houses of Worship"),
        ("parks",       "Parks"),
        ("health",      "Health & Social Services"),
    ]:
        df = assets.get(key, pd.DataFrame())
        if df.empty:
            continue
        available = [c for c in cols if c in df.columns]
        tbl = df[available].to_html(
            index=False, border=0, classes="",
            na_rep="—",
            float_format=lambda x: f"{x:.5f}",
        )
        sections.append(f"<h2>{label} ({len(df)} records)</h2>\n{tbl}")

    html = (
        f"<html><head><meta charset='utf-8'>"
        f"<title>Bronzeville Asset Summary</title>"
        f"<style>{css}</style></head><body>"
        f"<h1>Bronzeville Community Asset Map — Summary Tables</h1>"
        + "\n".join(sections)
        + "</body></html>"
    )

    with open(path, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"  Summary HTML  → {path}")
